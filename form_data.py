import csv
import logging_config
import os
import sys
import time
from datetime import datetime
from typing import Dict, List, Tuple

import pandas as pd
from openpyxl import Workbook
from pypdf import PdfReader
from tqdm import tqdm
from sklearn.model_selection import train_test_split

import call_api
import config
import logging

logger = logging.getLogger(__name__)

class DatasetMaker:
    def __init__(
        self,
        folder: str = config.DATA_DIR,
        xlsx_filename: str = config.XLSX_FILENAME,
        csv_filename: str = config.CSV_FILENAME_PAIRS,
        chunk_length_limit: int = config.CHUNK_LENGTH_LIMIT,
        chunk_overlap: int = config.CHUNK_OVERLAP,
    ):
        self.folder = folder
        self.xlsx_filename = xlsx_filename
        self.csv_filename = csv_filename
        self.chunk_length_limit = chunk_length_limit
        self.chunk_overlap = chunk_overlap


        logger.info("Инициализация DatasetMaker...")
        logger.info(f"Папка с данными: {folder}")
        logger.info(f"Лимит длины чанка: {chunk_length_limit}")
        logger.info(f"Перекрытие чанков: {chunk_overlap}")

        self.file_names = []
        self.file_contents = {}
        self.dataset = []

    def _load_data(self) -> None:
        self.file_names = self._get_data()
        self.file_contents = self._read_data()

    def _get_data(self) -> List[str] | None:
        try:
            files = [
                f
                for f in os.listdir(self.folder)
                if os.path.isfile(os.path.join(self.folder, f))
            ]
            if not files:
                logger.warning(f"В папке {self.folder} нет файлов")
            else:
                logger.info(f"Найдено {len(files)} файлов")
            return files

        except FileNotFoundError:
            logger.warning(f"Папка {self.folder} не найдена.")
            raise
        except Exception as e:
            logger.error(f"Ошибка при получении списка файлов: {e}")

    def _read_data(self) -> Dict[str, List[str]]:
        if not self.file_names:
            logger.error("Нет файлов для обработки.")
            raise ValueError("Нет файлов для обработки.")

        files_data = {}

        try:
            for file in self.file_names:
                reader = PdfReader(os.path.join(self.folder, file))
                chunks = []
                current_chunk = []
                for page in reader.pages:
                    text = page.extract_text()
                    words = text.split()
                    current_chunk = []
                    chunk_length = 0
                    for word in words:
                        chunk_length += len(word)
                        current_chunk.append(word)
                        if chunk_length > self.chunk_length_limit:
                            chunk = " ".join(current_chunk)
                            chunks.append(chunk)
                            previous_chunk = current_chunk[::-1]
                            current_chunk = []
                            chunk_length = 0

                            for previous_word in previous_chunk:
                                chunk_length += len(previous_word)
                                current_chunk.append(previous_word)
                                if chunk_length > self.chunk_overlap:
                                    current_chunk = current_chunk[::-1]
                                    break
                chunk = " ".join(current_chunk)
                chunks.append(chunk)
                files_data[file] = chunks
                logger.info(f"Создано {len(chunks)} чанков")
            return files_data
        except Exception as e:
            logger.error(f"Ошибка при чанкировании файлов: {e}")
            raise

    def _save_chunks_to_excel(self, data: Dict[str, List[str]]) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Чанки"
        ws.cell(row=1, column=1, value="Название файла")
        ws.cell(row=1, column=2, value="Фрагмент документа")
        current_row = 2

        for file, chunks in data.items():
            ws.cell(row=current_row, column=1, value=file)

            for chunk in chunks:
                ws.cell(row=current_row, column=2, value=chunk)
                current_row += 1

        wb.save(self.xlsx_filename)

    def _save_dataset_to_excel(self, data: List[str]) -> None:
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = config.XLSX_TITLE
            ws.cell(row=1, column=1, value=config.XLSX_QUESTION_ROW)
            ws.cell(row=1, column=2, value=config.XLSX_ANSWER_ROW)
            current_row = 2

            for fragment in data:
                question = fragment[: fragment.find(config.ANSWER_TAG)]
                answer = fragment[fragment.find(config.ANSWER_TAG) :]
                if question != answer:
                    question = question.replace(
                        config.QUESTION_TAG, config.QUESTION_REPLACEMENT
                    )
                    answer = answer.replace(
                        config.ANSWER_TAG, config.ANSWER_REPLACEMENT
                    )
                    ws.cell(row=current_row, column=1, value=question)
                    ws.cell(row=current_row, column=2, value=answer)
                else:
                    ws.cell(row=current_row, column=3, value=1)
                    ws.cell(row=current_row, column=2, value=question)
                current_row += 1
            logger.info(f"Датасет сохранён в {self.xlsx_filename}")
            wb.save(self.xlsx_filename)
        except Exception as e:
            logger.error(f"Ошибка при сохранении датасета в Excel: {e}")

    def _excel_to_csv(self, delimiter=config.CSV_DELIMITER) -> None:
        try:
            df = pd.read_excel(self.xlsx_filename)
            columns = [df.iloc[:, i].dropna().tolist() for i in range(df.shape[1])]

            for i in range(len(columns)):
                for j in range(len(columns[0])):
                    columns[i][j] = (
                        columns[i][j]
                        .replace(config.QUESTION_TAG, config.QUESTION_REPLACEMENT)
                        .replace(config.ANSWER_TAG, config.ANSWER_REPLACEMENT)
                        .rstrip()
                    )

            with open(self.csv_filename, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f, delimiter=delimiter)
                writer.writerow(config.CSV_P_FIELDNAMES)
                for i in range(len(columns[0])):
                    data = [columns[0][i], columns[1][i]]
                    writer.writerow(data)

            logger.info(f"Датасет преобразован в {self.csv_filename}")
        except Exception as e:
            logger.error(f"Ошибка при преобразовании данных из Excel в csv: {e}")

    @staticmethod
    def _read_csv(filename: str = config.CSV_FILENAME_TRIPLETS) -> List:
        strings = []
        with open(filename, "r", encoding="utf-8") as f:
            csv_reader = csv.reader(f, delimiter=config.CSV_DELIMITER)
            next(csv_reader)
            for row in csv_reader:
                strings.append(config.CSV_DELIMITER.join(r for r in row[:4]))

        return strings

    def _split_data_into_csv(self):
        strings = self._read_csv()

        n_groups = len(strings) // config.AMOUNT_OF_TRIPLETS_PER_QUESTION
        groups = list(range(n_groups))

        train_groups, test_groups = train_test_split(
            groups,
            train_size=config.TRAIN_SIZE / 100,
            test_size=config.TEST_SIZE / 100,
            random_state=42
        )

        train_data = []
        for i in train_groups:
            for j in range(3):
                row = strings[j + i * config.AMOUNT_OF_TRIPLETS_PER_QUESTION].split(config.CSV_DELIMITER)
                train_data.append(row)

        test_data = []
        for i in test_groups:
            for j in range(3):
                row = strings[j + i * config.AMOUNT_OF_TRIPLETS_PER_QUESTION].split(config.CSV_DELIMITER)
                test_data.append(row)

        columns = config.CSV_T_FIELDNAMES

        train_df = pd.DataFrame(train_data, columns=columns)
        test_df = pd.DataFrame(test_data, columns=columns)

        train_df = train_df.sample(frac=1).reset_index(drop=True)
        test_df = test_df.sample(frac=1).reset_index(drop=True)

        train_df.to_csv(
            config.CSV_TRAIN_FILENAME_TRIPLETS,
            index=False,
            header=True,
            sep=config.CSV_DELIMITER,
            quoting=csv.QUOTE_NONE,
            encoding="utf-8"
        )

        test_df.to_csv(
            config.CSV_TEST_FILENAME_TRIPLETS,
            index=False,
            header=True,
            sep=config.CSV_DELIMITER,
            quoting=csv.QUOTE_NONE,
            encoding="utf-8"
        )

    def run_parsing_pipeline(self) -> None:
        """
        Запускает пайплайн чтения данных, вызова локальной LLM для формирования пар вопрос-ответ
        и формирует выходные xlsx- и csv-файлы.
        """
        self._load_data()
        try:
            if not self.file_contents:
                logger.error("Нет данных для обработки")
                return
            for file, chunks in tqdm(self.file_contents.items()):
                logger.info(f"Начинаю работу над файлом {file}")
                for chunk in tqdm(chunks):
                    timeout_counts = 0
                    while True:
                        logger.info(
                            f"Отправляю чанк со следующим содержимым: {chunk}"
                        )
                        response = call_api.ask_model(chunk, config.DATASET_FORMING_PROMPT)
                        match response:
                            case "Timeout":
                                logger.warning(
                                    f"Превышено время ожидания. "
                                    f"Ожидаю 30 секунд перед повторной попыткой. "
                                    f"Количество попыток: {timeout_counts + 1}"
                                )
                                timeout_counts += 1
                                if timeout_counts >= 3:
                                    timeout_counts = 0
                                    break
                                time.sleep(30)
                                continue
                            case "Error":
                                logger.warning(
                                    "Ошибка на сервере. Ожидаю 60 секунд перед повторной попыткой."
                                )
                                time.sleep(60)
                                continue
                            case _:
                                logger.info("Успешно.")
                                for r in response.split("\n"):
                                    if len(r) > 10:
                                        self.dataset.append(r)
                                break

            self._save_dataset_to_excel(self.dataset)
            self._excel_to_csv()
            logger.info(f"Выполнение пайплайна парсинга завершено.")

        except KeyboardInterrupt:
            logger.warning("Обработка прервана пользователем")
            if self.dataset:
                logger.info("Сохранение частичных результатов...")
                self._save_dataset_to_excel(self.dataset)
                self._excel_to_csv()
        except Exception as e:
            logger.error(f"Ошибка при выполнении пайплайна: {e}")

    def run_triplet_pipeline(self):
        self._split_data_into_csv()